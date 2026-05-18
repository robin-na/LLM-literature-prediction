import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from utils.consensus import load_human_datasets
from utils.csvio import read_csv
from utils.paths import resolve_input_dataset_path


def _write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    path.write_text(
        "\n".join([",".join(header)] + [",".join(row) for row in rows]) + "\n",
        encoding="utf-8",
    )


def _write_xlsx(path: Path, header: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


class TestInputFormats(unittest.TestCase):
    def test_read_csv_supports_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = Path(tmp) / "sample.xlsx"
            _write_xlsx(
                xlsx_path,
                ["Filename", "Empirical"],
                [["paper1", "1"], ["paper2", "0"]],
            )

            rows = read_csv(xlsx_path)

            self.assertEqual(
                rows,
                [
                    {"Filename": "paper1", "Empirical": "1"},
                    {"Filename": "paper2", "Empirical": "0"},
                ],
            )

    def test_resolve_input_dataset_path_prefers_newer_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            directory = Path(tmp)
            csv_path = directory / "human_generated.csv"
            xlsx_path = directory / "human_generated.xlsx"
            _write_csv(csv_path, ["Filename"], [["paper_csv"]])
            _write_xlsx(xlsx_path, ["Filename"], [["paper_xlsx"]])

            resolved = resolve_input_dataset_path(directory, "human_generated")

            self.assertEqual(resolved, xlsx_path)

    def test_load_human_datasets_reads_xlsx_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            directory = Path(tmp)
            _write_xlsx(
                directory / "human_generated.xlsx",
                ["Filename", "Empirical"],
                [["paper1", "1"]],
            )

            datasets = load_human_datasets(directory)

            self.assertEqual(list(datasets.keys()), ["human_generated"])
            self.assertEqual(datasets["human_generated"][0]["Filename"], "paper1")

    def test_read_csv_prefers_data_sheet_over_active_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = Path(tmp) / "LLM_generated.xlsx"
            workbook = Workbook()
            empty_sheet = workbook.active
            empty_sheet.title = "Empty Table"
            empty_sheet.append(["Filename", "Title"])
            empty_sheet.append(["paper1", "wrong"])
            data_sheet = workbook.create_sheet("LLM-Generated")
            data_sheet.append(["custom_id", "METHOD_empirical"])
            data_sheet.append(["paper1.md", "TRUE"])
            workbook.save(xlsx_path)
            workbook.close()

            rows = read_csv(xlsx_path)

            self.assertEqual(rows[0]["custom_id"], "paper1.md")
