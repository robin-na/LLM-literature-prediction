from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


Row = dict[str, str]


def read_csv(path: Path) -> list[Row]:
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx(path)
    with path.open("r", encoding="utf-8", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def write_csv(path: Path, rows: Iterable[Row]) -> int:
    rows_list = list(rows)
    if not rows_list:
        return 0

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows_list[0].keys()))
        writer.writeheader()
        writer.writerows(rows_list)
    return len(rows_list)


def _read_xlsx(path: Path) -> list[Row]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = _select_worksheet(workbook)
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = ["" if value is None else str(value) for value in header_row]
        out: list[Row] = []
        for values in rows_iter:
            row: Row = {}
            has_value = False
            for idx, header in enumerate(headers):
                cell_value = values[idx] if idx < len(values) else None
                if cell_value not in {None, ""}:
                    has_value = True
                row[header] = "" if cell_value is None else str(cell_value)
            if has_value:
                out.append(row)
        return out
    finally:
        workbook.close()


def _select_worksheet(workbook) -> object:
    best_sheet = workbook.active
    best_score = float("-inf")
    for sheet in workbook.worksheets:
        header_row = next(sheet.iter_rows(values_only=True), None)
        score = _worksheet_score(sheet.title, header_row)
        if score > best_score:
            best_score = score
            best_sheet = sheet
    return best_sheet


def _worksheet_score(title: str, header_row: tuple[object, ...] | None) -> float:
    if not header_row:
        return float("-inf")

    headers = ["" if value is None else str(value).strip() for value in header_row]
    header_set = {header for header in headers if header}
    title_norm = title.strip().lower()

    score = float(len(header_set))
    if "custom_id" in header_set:
        score += 200.0
    if "Filename" in header_set:
        score += 120.0
    if "paper_id" in header_set:
        score += 80.0
    if "review_session_id" in header_set:
        score += 40.0
    if "consensus_session_id" in header_set:
        score += 40.0
    if "llm" in title_norm:
        score += 25.0
    if "human" in title_norm:
        score += 20.0
    if "empty" in title_norm:
        score -= 100.0
    return score

